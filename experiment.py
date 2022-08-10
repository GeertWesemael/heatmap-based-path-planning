import pickle

import numpy as np
import random
import heatmap
import map_
import path
import actor
import prob_heatmap
import robot
import world
import astar
import zone
import createactors
import matplotlib.pyplot as plt
from timefunct import sec_to_hour, hour_min_to_sec
from timefunct import random_time_between, random_time_between_, sec_to_hour_min_string,unifromly_dist_time_
from openpyxl import load_workbook
from datetime import datetime

for actors_ in [10,30,
                hour_min_to_sec(0, 1),hour_min_to_sec(0, 2),
                hour_min_to_sec(0, 5),hour_min_to_sec(0, 10),hour_min_to_sec(0, 20),hour_min_to_sec(0, 30),
                hour_min_to_sec(0, 40),hour_min_to_sec(0, 50),hour_min_to_sec(1, 0),hour_min_to_sec(2, 0)]:

    ##################################

    val = 1
    experiment = "prob_tf"
    amount_of_robots = 10

    if val == 1:
        ########### easy setup ###########
        robot_start_time = hour_min_to_sec(10, 0)
        interval = actors_
        start = hour_min_to_sec(9, 30)
        end = hour_min_to_sec(11, 0)
        start_loc_robot = (6, 7)
        end_loc_robot = (6, 0)

    if val == 2 or val == 3:
        ########### hard setup ###########
        robot_start_time = hour_min_to_sec(10, 0)
        interval = actors_
        start = hour_min_to_sec(9, 30)
        end = hour_min_to_sec(10, 30)
        start_loc_robot = (0, 6)
        end_loc_robot = (8, 6)

    ###################################
    # adjustable param
    a_star_weight_factor = 90
    amount_of_sample_worlds = 90
    amount_of_test_worlds = 10

    ###################################
    # standard parameters (don't touch)
    sample_r_col = 0.1
    sample_r_wait = 0.1
    distance_robot = 2
    wait_time = 2
    ###################################

    filename = ''
    if val == 1:
        print("Run easy setup")
        filename = 'easy_scenario_worlds_20'

    if val == 2:
        print("Run first hard setup")
        filename = 'hard_scenario_worlds_1_20'

    if val == 3:
        print("Run second hard setup")
        filename = 'hard_scenario_worlds_2_20'

    if filename == '':
        print("No correct value given")
    # Unpickle
    infile = open(filename, 'rb')
    list_of_worlds = pickle.load(infile)
    infile.close()
    print("Unpickled")

    # read all test worlds
    world1 = list_of_worlds[0]
    world_test = list_of_worlds[amount_of_sample_worlds + amount_of_test_worlds - 1]
    world_tests = list_of_worlds[amount_of_sample_worlds:amount_of_sample_worlds + amount_of_test_worlds]
    list_of_worlds = list_of_worlds[0:amount_of_sample_worlds]
    print(f"number of worlds used: {len(list_of_worlds)} / number of test worlds: {len(world_tests)}")

    print("plotting the world")
    # world1.plot_world()


    coll = None
    time = None
    r1 = None
    heatm = None
    heatmaps = None
    prob_heatmaps = None
    path_taken = None

    if experiment == "total":
        print("total heatmap")
        heatm = heatmap.Heatmap(world1, sample_rate=1, scale=1)
        heatm.visualize_heatmap()

    if experiment == "tf":
        print("moment heatmap")
        heatmaps = heatmap.heatmap_for_each_interval(world1, interval, start_time=start, end_time=end, sample_rate=1,
                                                     scale=1)
        # heatmap.animate_heatmaps(heatmaps)

    if experiment == "prob_tf":
        print("moment prob_heatmap")
        prob_heatmaps = prob_heatmap.heatmap_for_each_interval(list_of_worlds, interval, start_time=start, end_time=end,
                                                               sample_rate=1, scale=1)
        # prob_heatmap.animate_heatmaps(prob_heatmaps)


    #################################################################################
    for value in range(40, 500):
        print("start value " + str(value))
        a_star_weight_factor = value

        print("finding robot paths")
        map1 = world1.get_map()

        if amount_of_robots == 1:
            if experiment == "total":
                r1 = robot.Robot(start_loc_robot, robot_start_time, map1)
                r1.weighted_astar_path_plan(end_loc_robot, heatm, a_star_weight_factor)
                coll = r1.evaluate_collisions_worlds(world_tests, sample_r_col)
                time = r1.evaluate_time()
                # r1.path.plot_path(map1, "r1 10h global heatmap")

            if experiment == "tf":
                r1 = robot.Robot(start_loc_robot, robot_start_time, map1)
                r1.weighted_astar_path_plan_timeframes(end_loc_robot, heatmaps, a_star_weight_factor)
                coll = r1.evaluate_collisions_worlds(world_tests, sample_r_col)
                time = r1.evaluate_time()
                # r1.path.plot_path(map1, "r2 10h timeframed heatmap")

            if experiment == "prob_tf":
                r1 = robot.Robot(start_loc_robot, robot_start_time, map1)
                r1.weighted_astar_path_plan_timeframes(end_loc_robot, prob_heatmaps, a_star_weight_factor)
                coll = r1.evaluate_collisions_worlds(world_tests, sample_r_col)
                time = r1.evaluate_time()
                # r1.path.plot_path(map1, "r3 10h prob timeframed heatmap")

            if experiment == "wait":
                r1 = robot.Robot(start_loc_robot, robot_start_time, map1)
                r1.plan_path_waiting_at_encounter(world1, end_loc_robot, sample_r_wait, distance_robot, wait_time)
                coll = r1.evaluate_collisions_worlds(world_tests, sample_r_col)
                time = r1.evaluate_time()
                # r1.path.plot_path(map1, "r4 10h wait")

            path_taken = str(r1.path.path_list)
        else:
            path_taken = ""
            coll = 0
            time = 0
            start_times = unifromly_dist_time_(start, end, amount_of_robots)
            for n in range(amount_of_robots):
                robot_start_time = start_times[n]

                if experiment == "total":
                    r1 = robot.Robot(start_loc_robot, robot_start_time, map1)
                    r1.weighted_astar_path_plan(end_loc_robot, heatm, a_star_weight_factor)
                    coll += r1.evaluate_collisions_worlds(world_tests, sample_r_col)
                    time += r1.evaluate_time()
                    #r1.path.plot_path(map1, "r1 10h global heatmap")

                if experiment == "tf":
                    r1 = robot.Robot(start_loc_robot, robot_start_time, map1)
                    r1.weighted_astar_path_plan_timeframes(end_loc_robot, heatmaps, a_star_weight_factor)
                    coll += r1.evaluate_collisions_worlds(world_tests, sample_r_col)
                    time += r1.evaluate_time()
                    #r1.path.plot_path(map1, "r2 10h timeframed heatmap")

                if experiment == "prob_tf":
                    r1 = robot.Robot(start_loc_robot, robot_start_time, map1)
                    r1.weighted_astar_path_plan_timeframes(end_loc_robot, prob_heatmaps, a_star_weight_factor)
                    coll += r1.evaluate_collisions_worlds(world_tests, sample_r_col)
                    time += r1.evaluate_time()
                    #r1.path.plot_path(map1, "r3 10h prob timeframed heatmap")

                if experiment == "wait":
                    r1 = robot.Robot(start_loc_robot, robot_start_time, map1)
                    r1.plan_path_waiting_at_encounter(world1, end_loc_robot, sample_r_wait, distance_robot, wait_time)
                    coll += r1.evaluate_collisions_worlds(world_tests, sample_r_col)
                    time += r1.evaluate_time()
                    #r1.path.plot_path(map1, "r4 10h wait")`
                path_taken += str(r1.path.path_list)
            coll = coll/amount_of_robots
            time = time/amount_of_robots
            robot_start_time = 0

        # Write data to TXT file
        # Open a file with access mode 'a'
        file_object = open('results.txt', 'a')
        # Append 'hello' at the end of file
        file_object.write(f'date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, world: {val}, start: {sec_to_hour_min_string(start)}, end: {sec_to_hour_min_string(end)}, '
                          f'interval: {sec_to_hour_min_string(interval)}, r_start: {sec_to_hour_min_string(robot_start_time)}, '
                          f'wf: {a_star_weight_factor}, s_w: {amount_of_sample_worlds}, t_w: {amount_of_test_worlds}, '
                          f'exp: {experiment}, coll: {coll}, time: {time}, filename: {filename}, robots: {amount_of_robots}')
        file_object.write('\n')
        # Close the file
        file_object.close()

        # eddit spreadsheet
        workbook = load_workbook(filename="results.xlsx")

        sheet = workbook.active
        new_row = sheet.max_row + 1

        sheet.cell(row=new_row, column=1).value = val
        sheet.cell(row=new_row, column=2).value = sec_to_hour_min_string(start)
        sheet.cell(row=new_row, column=3).value = sec_to_hour_min_string(end)
        sheet.cell(row=new_row, column=4).value = sec_to_hour_min_string(interval)
        sheet.cell(row=new_row, column=5).value = sec_to_hour_min_string(robot_start_time)
        sheet.cell(row=new_row, column=6).value = a_star_weight_factor
        sheet.cell(row=new_row, column=7).value = amount_of_sample_worlds
        sheet.cell(row=new_row, column=8).value = amount_of_test_worlds
        sheet.cell(row=new_row, column=9).value = experiment
        sheet.cell(row=new_row, column=10).value = coll
        sheet.cell(row=new_row, column=11).value = time
        sheet.cell(row=new_row, column=12).value = path_taken
        sheet.cell(row=new_row, column=13).value = filename
        sheet.cell(row=new_row, column=14).value = amount_of_robots

        workbook.save(filename="results.xlsx")

        if coll == 0:
            break